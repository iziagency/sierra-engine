"""The CA filing questions have to answer for a real dossier.

Lakeside is a California risk with CA filing number 564061 and an MCP-65 on
file, and the RTS came out with "CA State filing required?" and "Filing State"
blank. The cause: rts_fill read `company.state`, and no real dossier carries
that key — the state exists only inside the address string.

`routing._state` already had to solve this to apply the CA half of JC's routing
rule, so both readers use the same one.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

import rts_fill

DOSSIER = {
    "sp_code": "LAKES",
    "company": {
        "first_named_insured": "Lakeside Towing LLC",
        "entity_type": "llc",
        "state_filing_number": "564061",
        "location_address": "7050 Perris Hill Rd, San Bernardino, CA 92404",
        "mailing_address": "7050 Perris Hill Rd, San Bernardino, CA 92404",
    },
    "coverages": {"auto_liability": "1m"},
    "operations": {"tow_disabled_autos": 90},
    "vehicles": [], "drivers": [], "loss_runs": [],
}


def fill(tmp_path, monkeypatch, dossier):
    clients_root = tmp_path / "clients"
    slug = "lakeside-towing-llc"
    (clients_root / slug).mkdir(parents=True)
    (clients_root / slug / "state.json").write_text(json.dumps(dossier),
                                                    encoding="utf-8")
    monkeypatch.setattr(rts_fill, "CLIENTS", clients_root)
    monkeypatch.setattr(rts_fill, "OUT", tmp_path / "out")
    r = rts_fill.fill(slug)
    assert r["ok"] is True
    return load_workbook(r["file"])


def test_ca_filing_is_answered_from_the_address(tmp_path, monkeypatch):
    wb = fill(tmp_path, monkeypatch, DOSSIER)
    assert wb["Business"].cell(row=17, column=5).value == "Yes"


def test_the_filing_state_is_named(tmp_path, monkeypatch):
    wb = fill(tmp_path, monkeypatch, DOSSIER)
    assert wb["Business"].cell(row=33, column=4).value == "CA"


def test_an_explicit_state_field_still_wins_when_present(tmp_path, monkeypatch):
    d = json.loads(json.dumps(DOSSIER))
    d["company"]["state"] = "CA"
    wb = fill(tmp_path, monkeypatch, d)
    assert wb["Business"].cell(row=33, column=4).value == "CA"


def test_an_out_of_state_risk_is_not_told_it_needs_a_ca_filing(tmp_path, monkeypatch):
    d = json.loads(json.dumps(DOSSIER))
    d["company"]["location_address"] = "1 Main St, Dallas, TX 75201"
    d["company"]["mailing_address"] = "1 Main St, Dallas, TX 75201"
    wb = fill(tmp_path, monkeypatch, d)
    assert wb["Business"].cell(row=17, column=5).value is None
    assert wb["Business"].cell(row=33, column=4).value is None
