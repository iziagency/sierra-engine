"""Nothing leaves the RTS unmentioned.

The workbook holds 9 power units, 9 trailers and 20 drivers. The filler wrote
`vehicles[:9]` and `drivers[:20]` and said nothing about the rest, and never
touched the trailer section at all — so a fleet of 8 trucks and 3 trailers
passed the 1-9 routing rule (trailers are not power units), then quietly lost
two units on the way into the spreadsheet.

A submission that goes out short is worse than one that does not go out: the
underwriter quotes what is on the paper. Trailers now land in the trailer
section where they belong, and anything that genuinely will not fit is
reported, never dropped in silence.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

import rts_fill


def truck(i):
    return {"year": 2020 + (i % 5), "maker": "Ford", "body_type": "carrier",
            "vin": f"1FDUF5HT8PDA{i:05d}", "gvw": 19500, "stated_value": 50000 + i}


def trailer(i):
    return {"year": 2019, "maker": "Big Tex", "body_type": "trailer",
            "vin": f"5JWBU2029KP{i:06d}", "is_trailer": True,
            "stated_value": 12000 + i, "length": "20 ft"}


def build(tmp_path, monkeypatch, vehicles, drivers=()):
    clients = tmp_path / "clients"
    slug = "fleet-test-llc"
    (clients / slug).mkdir(parents=True)
    (clients / slug / "state.json").write_text(json.dumps({
        "sp_code": "FLEET1",
        "company": {"first_named_insured": "Fleet Test LLC",
                    "location_address": "1 Main St, Fresno, CA 93701"},
        "coverages": {}, "loss_runs": [],
        "vehicles": list(vehicles), "drivers": list(drivers),
    }), encoding="utf-8")
    monkeypatch.setattr(rts_fill, "CLIENTS", clients)
    monkeypatch.setattr(rts_fill, "OUT", tmp_path / "out")
    r = rts_fill.fill(slug)
    assert r["ok"] is True
    return r, load_workbook(r["file"])


def test_trailers_go_to_the_trailer_section_not_the_truck_columns(tmp_path, monkeypatch):
    r, wb = build(tmp_path, monkeypatch, [truck(1), truck(2), trailer(1)])
    veh = wb["Vehicles"]
    # trucks occupy D and E of the power-unit block (VIN row 13)
    assert veh.cell(row=13, column=4).value == truck(1)["vin"]
    assert veh.cell(row=13, column=5).value == truck(2)["vin"]
    assert veh.cell(row=13, column=6).value is None      # not a third truck
    # the trailer lands in its own block (VIN row 26)
    assert veh.cell(row=26, column=4).value == trailer(1)["vin"]


def test_eight_trucks_and_three_trailers_all_land(tmp_path, monkeypatch):
    # Passes the 1-9 routing rule on power units, used to lose two rows.
    vehicles = [truck(i) for i in range(1, 9)] + [trailer(i) for i in range(1, 4)]
    r, wb = build(tmp_path, monkeypatch, vehicles)
    veh = wb["Vehicles"]
    trucks_written = [veh.cell(row=13, column=c).value for c in range(4, 13)]
    assert len([v for v in trucks_written if v]) == 8
    trailers_written = [veh.cell(row=26, column=c).value for c in range(4, 13)]
    assert len([v for v in trailers_written if v]) == 3
    assert r["dropped"] == []


def test_a_fleet_too_big_for_the_sheet_is_reported_not_dropped(tmp_path, monkeypatch):
    r, _ = build(tmp_path, monkeypatch, [truck(i) for i in range(1, 12)])
    assert r["dropped"], "11 trucks into 9 columns must be reported"
    joined = " ".join(r["dropped"])
    assert "11" in joined and "9" in joined


def test_too_many_drivers_is_reported_too(tmp_path, monkeypatch):
    drivers = [{"name": f"Driver {i}", "state": "CA"} for i in range(1, 24)]
    r, _ = build(tmp_path, monkeypatch, [truck(1)], drivers)
    joined = " ".join(r["dropped"])
    assert "23" in joined and "20" in joined


def test_a_fleet_that_fits_reports_nothing(tmp_path, monkeypatch):
    r, _ = build(tmp_path, monkeypatch, [truck(1), truck(2)],
                 [{"name": "Jane Owner", "state": "CA"}])
    assert r["dropped"] == []
