"""The RTS is read by an underwriter, so it cannot leak our internals.

Found reviewing the two workbooks built for Allyssa's 7.31 assignment, from
JC's real QPs. Every defect is the same shape — a value that is fine inside the
engine reaching a page a human quotes from:

    Radius Of Operations      "100% lt50"            <- dossier keys
    Vehicle type              "Wl"                   <- unmapped abbreviation
    Business Owner Home addr  "See location schedule" <- a pointer, three times
    Business Name & DBA       "Wendy Hamilton"        <- the person, no business

And HAMIL's liability limit came out blank with nothing said about it, which
under JC's three-states rule is the one kind of blank that must speak up.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

import rts_fill

REAL_ADDR = "3615 Campbell St, Riverside, CA 92509"

HAMIL = {
    "sp_code": "HAMIL",
    "company": {
        "first_named_insured": "Wendy Hamilton",
        "dba": "Hamilton's Towing",
        "owner_name": "Wendy Hamilton",
        "entity_type": "sole_proprietor",
        "location_address": "See location schedule",
        "mailing_address": REAL_ADDR,
        "state_filing_number": "0151266",
    },
    "location": {"address": REAL_ADDR, "home_based": True},
    "coverages": {},
    "radius": {"lt50": 100},
    "vehicles": [{"year": 2015, "maker": "Chevrolet", "body_type": "Wl",
                  "vin": "1GC4CYEG0FF656041", "stated_value": 30000}],
    "drivers": [], "loss_runs": [],
}


@pytest.fixture
def book(tmp_path, monkeypatch):
    clients = tmp_path / "clients"
    slug = "hartley-towing"
    (clients / slug).mkdir(parents=True)
    (clients / slug / "state.json").write_text(json.dumps(HAMIL), encoding="utf-8")
    monkeypatch.setattr(rts_fill, "CLIENTS", clients)
    monkeypatch.setattr(rts_fill, "OUT", tmp_path / "out")
    r = rts_fill.fill(slug)
    assert r["ok"] is True
    return r, load_workbook(r["file"])


# ------------------------------------------------------- no internal keys

def test_the_radius_reads_in_english(book):
    _, wb = book
    assert wb["Start"].cell(row=26, column=4).value == "100% under 50 mi"


def test_every_radius_band_has_a_label():
    for key in ("lt50", "51_300", "301_500", "501_1000", "1000_2500",
                "2501_5000", "5000_plus"):
        assert key in rts_fill.RADIUS_LABELS, key
        assert "_" not in rts_fill.RADIUS_LABELS[key]


# ------------------------------------------------- no unmapped abbreviations

def test_a_body_type_abbreviation_is_spelled_out(book):
    _, wb = book
    assert wb["Vehicles"].cell(row=11, column=4).value == "Wheel Lift"


@pytest.mark.parametrize("raw,shown", [
    ("Wl", "Wheel Lift"), ("WL", "Wheel Lift"), ("wl", "Wheel Lift"),
    ("CC", "Car Carrier"), ("RB", "Car Carrier"), ("rollback", "Car Carrier"),
    ("wheel-lift", "Wheel Lift"), ("tractor", "Tractor"),
])
def test_the_shorthands_brokers_actually_type(raw, shown):
    assert rts_fill.body_label(raw) == shown


def test_a_body_type_we_do_not_know_is_passed_through_readably():
    assert rts_fill.body_label("service truck") == "Service Truck"


# ----------------------------------------------- no pointers on the document

def test_a_pointer_address_is_replaced_by_the_real_one(book):
    _, wb = book
    start = wb["Start"]
    assert start.cell(row=22, column=4).value == REAL_ADDR      # owner home
    assert REAL_ADDR in str(start.cell(row=27, column=4).value)  # lot location
    assert wb["Vehicles"].cell(row=14, column=4).value == REAL_ADDR


def test_see_location_schedule_appears_nowhere_in_the_workbook(book):
    _, wb = book
    for sh in wb.sheetnames:
        for row in wb[sh].iter_rows():
            for c in row:
                assert "location schedule" not in str(c.value or "").lower(), \
                    f"{sh}!{c.coordinate}"


# ------------------------------------------------ the business has to be named

def test_a_sole_proprietor_carries_their_dba(book):
    _, wb = book
    got = wb["Start"].cell(row=12, column=4).value
    assert "Wendy Hamilton" in got
    assert "Hamilton's Towing" in got


# ------------------------------------------------------- a hole has to speak

def test_a_missing_liability_limit_is_reported(book):
    r, wb = book
    assert wb["Business"].cell(row=13, column=4).value is None
    assert any("liability" in b.lower() for b in r["unknown_blanks"]), \
        r["unknown_blanks"]
