"""Wiring tests for reports/rts_fill.py: phone/date formatting and
cosmetic-dash stripping on the way into the RTS workbook, plus a regression
guard on the FEIN dash-INSERTION line, which must NOT change (flagged as an
open question for JC - see this change's final report).

The end-to-end test uses the real RTS template (reference/CAP RTS supp
app.xlsx, a blank form - no client data) but writes output only under
tmp_path; CLIENTS/OUT are monkeypatched so nothing under the real
app-form/clients tree is ever touched, per this change's hard constraints.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

import rts_fill

FIXTURE_DOSSIER = {
    "sp_code": "TESTT1",
    "company": {
        "first_named_insured": "Test Towing LLC",
        "entity_type": "llc",
        "fein": "85-0604178",
        "owner_name": "Jane Owner",
        "contact_cell": "909-685-9794",
        "office_phone": "(909) 685-9794",
        "state_filing_number": "CA-564061",
        "usdot_number": "3416380",
        "current_auto_carrier": "MSIG Specialty",
        "current_auto_expires": "8/3/2026",
        "mailing_address": "1 Main St, San Bernardino, CA 92404",
        "location_address": "1 Main St, San Bernardino, CA 92404",
        "total_vehicles": 1,
    },
    "coverages": {"auto_liability": "1m"},
    "vehicles": [],
    "drivers": [
        {"name": "Jane Owner", "state": "CA", "license": "B-5697020",
         "birthday": "8/31/1976"},
    ],
    "loss_runs": [],
}


class TestDriverClaims:
    def test_no_claims_at_all_is_a_confident_no(self):
        d = {"drivers": [{"name": "Jane Owner"}], "loss_runs": []}
        had, when, what = rts_fill._driver_claims(d, d["drivers"][0])
        assert (had, when, what) == ("No", "", "")

    def test_claim_date_is_formatted_mdyy(self):
        d = {
            "drivers": [{"name": "Jane Owner"}],
            "loss_runs": [{"claim_count": 1, "claims": [
                {"date_of_loss": "3/10/2025", "type": "Collision",
                 "status": "Closed", "total_incurred": 100.0,
                 "driver_name": "Owner"},
            ]}],
        }
        had, when, what = rts_fill._driver_claims(d, d["drivers"][0])
        assert had == "Yes"
        assert when == "3.10.25"

    def test_unparseable_claim_date_is_kept_not_dropped(self):
        d = {
            "drivers": [{"name": "Jane Owner"}],
            "loss_runs": [{"claim_count": 1, "claims": [
                {"date_of_loss": "unclear", "type": "Collision",
                 "status": "Closed", "total_incurred": 100.0,
                 "driver_name": "Owner"},
            ]}],
        }
        had, when, what = rts_fill._driver_claims(d, d["drivers"][0])
        assert when == "unclear"


class TestFillEndToEnd:
    """A real fill() run against the real RTS template, entirely inside
    tmp_path."""

    @pytest.fixture
    def client_slug(self, tmp_path, monkeypatch):
        clients_root = tmp_path / "clients"
        out_root = tmp_path / "out"
        slug = "test-towing-llc"
        cdir = clients_root / slug
        cdir.mkdir(parents=True)
        (cdir / "state.json").write_text(json.dumps(FIXTURE_DOSSIER), encoding="utf-8")
        monkeypatch.setattr(rts_fill, "CLIENTS", clients_root)
        monkeypatch.setattr(rts_fill, "OUT", out_root)
        return slug

    def test_phone_date_and_dash_strip_fields_in_the_written_workbook(self, client_slug):
        result = rts_fill.fill(client_slug)
        assert result["ok"] is True
        wb = load_workbook(result["file"])

        start = wb["Start"]
        assert start.cell(row=17, column=4).value == "CA564061"        # CA Authority # - dash stripped
        assert start.cell(row=21, column=4).value == "8.31.76"         # Business Owner Date of Birth
        assert start.cell(row=23, column=4).value == "909.685.9794"    # Business Owner primary phone#

        business = wb["Business"]
        assert business.cell(row=14, column=4).value == "8.3.26"       # Current Policy Expiration Date

        final = wb["Final Details"]
        # Explicitly UNCHANGED: the FEIN keeps its inserted dash (open
        # question for JC, not touched by this change).
        assert final.cell(row=21, column=4).value == "85-0604178"
        assert final.cell(row=38, column=4).value == "909.685.9794"    # Phone Number

        drivers = wb["Drivers"]
        assert drivers.cell(row=13, column=6).value == "8.31.76"       # birthday
        assert drivers.cell(row=13, column=7).value == "B5697020"      # license, dash stripped
