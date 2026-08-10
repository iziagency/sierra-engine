"""The RTS cookie-cutters JC dictated cell by cell on the 7.29 call.

Two of them REVERSE what he said on 7.22, which is why they are pinned here
rather than left to the older comments in rts_fill.py:

  * axles — 7.22 said "typically two unless it's GVW higher"; 7.29 said
    "I don't think we need to have how many axles" -> blank.
  * comp/collision — was written as a flat Yes; he said it is derived from
    whether the stated value is non-zero.

The rest are verbatim: Stetson every file, RT Specialty and the code never
change, the retail agency is the full legal name, spouse no, blanket waiver of
subrogation no, anyone-else-authorised no, loan/lease deliberately blank.

Blank is not one thing here. His rule: "every single answer has to have either
an answer accustomed to that insured, a cookie cutter answer, or we're
purposely leaving it blank because of some reason." A cell left blank because
nobody knows the number is a different object from a cell left blank on
purpose, and the output has to keep them apart.
"""
from __future__ import annotations

import json

import pytest
from openpyxl import load_workbook

import rts_fill

DOSSIER = {
    "sp_code": "TESTT1",
    "company": {
        "first_named_insured": "Test Towing LLC",
        "entity_type": "llc",
        "state": "CA",
        "location_address": "1 Main St, San Bernardino, CA 92404",
        "mailing_address": "1 Main St, San Bernardino, CA 92404",
        "total_vehicles": 3,
    },
    "coverages": {"auto_liability": "1m", "on_hook": "150k"},
    "vehicles": [
        # priced -> comp/collision Yes
        {"year": 2023, "maker": "Ford", "model": "F-550", "body_type": "carrier",
         "vin": "1FDUF5HT8PDA00001", "gvw": 19500, "stated_value": 98000},
        # explicitly worth nothing to insure -> No
        {"year": 2005, "maker": "Ford", "model": "F-350", "body_type": "wheel-lift",
         "vin": "1FDWF36P05EA00002", "gvw": 11500, "stated_value": 0},
        # nobody supplied a value -> unknown, which is NOT the same as No
        {"year": 2018, "maker": "Freightliner", "model": "M2", "body_type": "carrier",
         "vin": "1FVACWDT8JHJA0003", "gvw": 33000},
    ],
    "drivers": [],
    "loss_runs": [],
}


@pytest.fixture
def book(tmp_path, monkeypatch):
    clients_root = tmp_path / "clients"
    slug = "test-towing-llc"
    cdir = clients_root / slug
    cdir.mkdir(parents=True)
    (cdir / "state.json").write_text(json.dumps(DOSSIER), encoding="utf-8")
    monkeypatch.setattr(rts_fill, "CLIENTS", clients_root)
    monkeypatch.setattr(rts_fill, "OUT", tmp_path / "out")
    result = rts_fill.fill(slug)
    assert result["ok"] is True
    return load_workbook(result["file"])


# ------------------------------------------------------- Final Details tab

def test_retail_agency_is_the_full_legal_name(book):
    # "Sierra Pacific Insurance Services Inc — retail agency name, every time"
    assert book["Final Details"].cell(row=15, column=4).value == \
        "Sierra Pacific Insurance Services Inc"


def test_wholesaler_and_agent_code_never_change(book):
    final = book["Final Details"]
    assert final.cell(row=11, column=4).value == "RT Specialty"
    assert final.cell(row=14, column=4).value == "94767"


def test_spouse_is_always_no(book):
    assert book["Final Details"].cell(row=18, column=4).value == "No"


def test_blanket_waiver_of_subrogation_is_no(book):
    assert book["Final Details"].cell(row=34, column=4).value == "No"


def test_nobody_else_is_authorised_to_update_the_policy(book):
    assert book["Final Details"].cell(row=39, column=4).value == "No"


def test_no_access_level_is_granted_below_it(book):
    # Nobody is authorised, so there is no action to grant. The cell keeps the
    # template's own option hint — RTS ships that text; the engine must not
    # turn it into an answer by picking one.
    assert book["Final Details"].cell(row=40, column=4).value == \
        "(View Only Access, Edit Access, Full Access)"


def test_the_premium_finance_cookie_cutter_survives(book):
    # Stetson is "every file, never changes". RTS ships it pre-filled in the
    # template; the engine must not blank the tab it never writes to.
    pay = book["Payment"]
    assert pay.cell(row=11, column=4).value.strip() == "Stetson Insurance Funding, LLC"
    assert "Hunt Valley" in pay.cell(row=12, column=4).value


# ------------------------------------------------------------ Vehicles tab

def test_axles_is_blank_now(book):
    # 7.29 supersedes 7.22: "I don't think we need to have how many axles"
    veh = book["Vehicles"]
    for col in (4, 5, 6):
        assert veh.cell(row=15, column=col).value is None


def test_loan_lease_is_blank_on_purpose(book):
    # Answering yes makes the underwriter demand the loss-payee detail before
    # quoting, so it is left blank deliberately.
    veh = book["Vehicles"]
    for col in (4, 5, 6):
        assert veh.cell(row=17, column=col).value is None


def test_comp_collision_follows_the_stated_value(book):
    veh = book["Vehicles"]
    assert veh.cell(row=20, column=4).value == "Yes"     # 98,000
    assert veh.cell(row=20, column=5).value == "No"      # explicitly 0


def test_a_missing_stated_value_leaves_comp_collision_blank(book):
    # Unknown is not No. Guessing No here silently drops physical damage
    # coverage from the quote.
    assert book["Vehicles"].cell(row=20, column=6).value is None


# --------------------------------------------------------------- money case

def test_the_liability_limit_prints_a_capital_m(book):
    # "$1M with a capital M, not lowercase"
    assert book["Business"].cell(row=13, column=4).value == "$1M"
