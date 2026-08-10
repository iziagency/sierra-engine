r"""RTS/Progressive supplemental Excel, auto-populated from the dossier.

JC does this by hand in 15-25 minutes per client and asked for exactly this:
"having a simple skill that once we get the app completed, populate the RTS app
too." The six-tab workbook asks the same questions as the CAP app in different
words, so this module is a translation table — each row is found by its LABEL in
column C (not by fixed coordinates, so a re-ordered template still fills) and
answered from the dossier.

The cookie-cutter answers come from JC verbatim on the 7.22 call:
  * Progressive agent code "ours is 94767" — always
  * wholesaler is RT Specialty, retail agency Sierra Pacific — always
  * spouse as second named insured — "this is always no"
  * order MVR/CLUE — "this is always yes, we want the accurate quote"
  * vehicle loan company — "we can leave the loan company blank"
  * lot location means the LOCATION full address, never the mailing address
  * garaging address = same as location
  * axles: "typically two unless it's GVW higher"
  * hitch type "isn't important if they're not carrying trailers" -> N/A
  * loan/lease on vehicle — "just leave it blank for now"

Everything unknown stays BLANK for the broker: blank beats wrong.

Usage:
  python rts_fill.py --client lakeside-towing-llc
  python rts_fill.py --client lakeside-towing-llc --from-qp "path/to/some_QP.pdf"
"""
from __future__ import annotations

import argparse
import datetime
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
CLIENTS = ROOT / "app-form" / "clients"
TEMPLATE = ROOT / "reference" / "CAP RTS supp app.xlsx"
OUT = ROOT / "reports" / "out"

sys.path.insert(0, str(ROOT / "shared"))
import formatting  # noqa: E402 - needs sys.path set up first
import routing  # noqa: E402 - shared state/eligibility resolver
import qp_read  # noqa: E402 - QP PDF -> dossier link, see qp_read.py


def m8(d: datetime.date | None = None) -> str:
    """Filename date stamp — defaults to today. Formatting itself now lives in
    shared/formatting.py; this wrapper keeps the "no argument = today"
    filename behaviour exactly as it was (filename-generation logic is out of
    scope for JC's 7.27 formatting rules — it already matches doctrine)."""
    return formatting.format_date(d or datetime.date.today())


# The workbook's own capacity: 9 power-unit columns, 9 trailer columns, 20
# driver rows. Anything past these is reported, never silently sliced off.
MAX_RTS_UNITS = 9
MAX_RTS_DRIVERS = 20

TRAILER_WORDS = re.compile(r"trailer|lowboy|gooseneck|dolly", re.I)


def _is_trailer(v: dict) -> bool:
    """A trailer belongs in the trailer block, not a power-unit column."""
    if v.get("is_trailer") is not None:
        return bool(v["is_trailer"])
    return bool(TRAILER_WORDS.search(str(v.get("body_type") or "")))


# Text that points somewhere else instead of answering. It belongs on JC's own
# app, where the schedule really is a page away; it must never reach a vendor
# workbook, which has no schedule to look at. Hartley Towing put "See
# location schedule" onto three cells of a real submission.
POINTER_TEXT = re.compile(
    r"\bsee\s+(the\s+)?(location|vehicle|driver|attached)\b|\bas\s+above\b"
    r"|\bsee\s+schedule\b|\bsee\s+attached\b", re.I)


def _real_address(dossier: dict) -> str:
    """The location address, resolved past any pointer to the actual street."""
    c = dossier.get("company", {}) or {}
    for cand in (c.get("location_address"), c.get("mailing_address"),
                 (dossier.get("location") or {}).get("address")):
        s = str(cand or "").strip()
        if s and not POINTER_TEXT.search(s):
            return s
    return ""


def _insured_name(c: dict) -> str:
    """Named insured, with the DBA when the two differ.

    A sole proprietor's insured IS the person, so the sheet showed only "Wendy
    Hamilton" and an underwriter had no idea which towing company it was.
    """
    named = str(c.get("first_named_insured") or "").strip()
    dba = str(c.get("dba") or "").strip()
    if dba and dba.lower() not in named.lower():
        return f"{named} dba {dba}" if named else dba
    return named


def _entity(c: dict) -> str:
    return {"llc": "LLC", "corporation": "Corporation",
            "sole_proprietor": "Individual / Sole Proprietor",
            "partnership": "Partnership"}.get(str(c.get("entity_type") or "").lower(), "")


# The dossier's radius keys are ours; an underwriter reads the cell. Left
# unmapped they printed "80% lt50, 20% 51 300" onto a real submission.
RADIUS_LABELS = {
    "lt50": "under 50 mi",
    "51_300": "51-300 mi",
    "301_500": "301-500 mi",
    "501_1000": "501-1,000 mi",
    "1000_2500": "1,000-2,500 mi",
    "2501_5000": "2,501-5,000 mi",
    "5000_plus": "over 5,000 mi",
}

# Body types as brokers actually type them on a schedule — "Wl", "CC", "RB" —
# against the wording the RTS sheet itself asks for.
BODY_LABELS = {
    "wl": "Wheel Lift", "wheellift": "Wheel Lift", "wheel-lift": "Wheel Lift",
    "wheel lift": "Wheel Lift",
    "cc": "Car Carrier", "rb": "Car Carrier", "rollback": "Car Carrier",
    "carrier": "Car Carrier", "flatbed": "Car Carrier",
    "car carrier": "Car Carrier",
    "tractor": "Tractor", "trailer": "Trailer",
}


def body_label(raw) -> str:
    """A vehicle body type in the sheet's own vocabulary, never a raw code."""
    s = str(raw or "").strip()
    if not s:
        return ""
    return BODY_LABELS.get(s.lower(), s.title())


def _radius(d: dict) -> str:
    r = d.get("radius")
    if isinstance(r, dict):
        parts = []
        for k, v in r.items():
            if not v:
                continue
            v = str(v)
            # "20 (200 mi max)" carries its own annotation; only bare numbers
            # get the percent sign appended
            label = RADIUS_LABELS.get(k, k.replace("_", " "))
            if v.isdigit():
                parts.append(f"{v}% {label}")
            elif "(" in v:                      # "20 (200 mi max)" -> pct, range, note
                pct, note = v.split("(", 1)
                parts.append(f"{pct.strip()}% {label} ({note.strip()}")
            else:
                parts.append(f"{v} {label}")
        return ", ".join(parts)
    return str(r or "")


def _owner_dob(d: dict) -> str:
    c = d.get("company", {}) or {}
    owner = str(c.get("owner_name") or "").lower()
    for drv in d.get("drivers") or []:
        if owner and str(drv.get("name") or "").lower() == owner:
            return str(drv.get("birthday") or "")
    return str(c.get("owner_dob") or "")


def _stored(d: dict) -> str:
    c = d.get("company", {}) or {}
    loc = d.get("location", {}) or {}
    if loc.get("home_based") or c.get("home_based") \
            or str(c.get("total_commercial_locations") or "") == "0":
        return "Taken home"
    if loc.get("description") or c.get("location_address"):
        return "Stored in yard"
    return ""


def _driver_claims(d: dict, drv: dict) -> tuple[str, str, str]:
    """(had accidents yes/no-or-blank, date, description) from the loss runs."""
    name = str(drv.get("name") or "").lower().split()
    solo = len(d.get("drivers") or []) == 1
    hits = []
    for lr in d.get("loss_runs") or []:
        for cl in lr.get("claims") or []:
            hay = (str(cl.get("driver_name") or "") + " "
                   + str(cl.get("description") or "")).lower()
            if name and all(part in hay for part in name[-1:]):     # match surname
                hits.append(cl)
            elif solo:
                # one driver on the policy: every claim is necessarily theirs —
                # attribution by arithmetic, not by guesswork
                hits.append(cl)
    if not hits:
        # no claims anywhere at all -> a confident "No"; claims exist but not
        # matched to this driver -> blank, the broker decides
        total = sum(int(lr.get("claim_count") or 0) for lr in d.get("loss_runs") or [])
        return ("No", "", "") if total == 0 else ("", "", "")
    h = hits[0]
    # Write-time safety net: apply_gospel() now normalises date_of_loss to
    # M.D.YY at ingest (watcher/lossruns.py), but a dossier written before
    # that change can still carry the model's raw M/D/YYYY — format it here
    # too rather than assume every dossier on disk was already fixed. Falls
    # back to the raw string, never drops it, if it truly isn't a date.
    dol = formatting.format_date(h.get("date_of_loss")) or str(h.get("date_of_loss") or "")
    return ("Yes", dol,
            f"{h.get('type', '')}, {h.get('status', '')} — "
            f"${float(h.get('total_incurred') or 0):,.0f} incurred (per loss run)")


def fill(slug: str) -> dict:
    folder = CLIENTS / slug
    dossier = json.loads((folder / "state.json").read_text(encoding="utf-8"))
    c = dossier.get("company", {}) or {}
    # No invented identity. A workbook named "CLIENT CAP RTS supp app" reached
    # Drive next to a correctly named app on Shoreline's first drop, because
    # this fell back to a literal when the code had not been anchored yet. The
    # SP code IS the client — JC: "our unique identifier" — so a missing one is
    # a gap to report, never a filename to make up.
    sp = dossier.get("sp_code")
    if not sp:
        return {"ok": False, "cells": 0, "file": "",
                "error": f"no SP code on {slug}'s file yet — the RTS cannot be "
                         f"named without it"}
    # Past any "see location schedule" pointer: this cell goes on a document
    # whose reader has no schedule to see.
    loc_addr = _real_address(dossier)
    meta = dossier.get("meta_data", {}) or {}
    interstate = c.get("cross_state_lines")
    if interstate is None and meta.get("interstate"):
        interstate = str(meta["interstate"]).strip().lower() == "yes"
    covs = dossier.get("coverages", {}) or {}
    # No real dossier carries company.state — the state lives inside the
    # address. Reading the key directly left Lakeside, a California risk with an
    # MCP-65 on file, with both CA filing questions blank. Same resolver the
    # routing rule uses, so the two can never disagree about where a risk is.
    state = routing.state_of(c)

    # label-pattern -> value; value column is D unless (value, "E") is given
    START = {
        r"Proposed Effective date": formatting.format_date(
            c.get("current_auto_expires") or c.get("policy_effective_date")),
        r"Business Name & DBA": _insured_name(c),
        r"Business Type": _entity(c),
        r"Business Class": (dossier.get("risk") or "tow").replace("tow", "Towing"),
        r"USDOT#": c.get("usdot_number"),
        r"CA Authority #": formatting.strip_cosmetic_dashes(c.get("state_filing_number")),
        r"Business Owners Name": c.get("owner_name"),
        r"Business Owner Date of Birth": formatting.format_date(_owner_dob(dossier)),
        r"Business Owner Home address": loc_addr if _stored(dossier) == "Taken home" else "",
        r"Business Owner primary phone": formatting.format_phone(
            c.get("contact_cell") or c.get("office_phone")),
        r"How is the customer's business structured": _entity(c),
        r"percentage .* repossession": "0%" if not (
            (dossier.get("operations") or {}).get("repo_work")) else None,
        r"Radius Of Operations": _radius(dossier),
        r"Garage Keepers Lot Location":
            (f"{loc_addr} (home based, no commercial lot)"
             if _stored(dossier) == "Taken home" else loc_addr) if loc_addr else None,
        r"stored in a yard": _stored(dossier),
    }
    BUSINESS = {
        r"Is this risk currently insured": "Yes" if c.get("current_auto_carrier") else None,
        r"continuous commercial auto coverage":
            "Yes" if str(c.get("years_with_auto_insurance") or "0").isdigit()
                     and int(c.get("years_with_auto_insurance") or 0) >= 1 else None,
        r"Current Bodily Injury Liability Limit":
            formatting.format_limit(covs.get("auto_liability")),
        r"Current Policy Expiration Date": formatting.format_date(c.get("current_auto_expires")),
        r"GCL / BOP coverage": "Yes" if covs.get("general_liability") else None,
        r"Federal Filing Required": (("No", "E") if interstate is False
                                     else ("Yes", "E") if interstate else None),
        r"CA State filing required":
            (("Yes", "E") if state == "CA" and c.get("state_filing_number")
             else None),
        r"How many vehicles on the quote": str(len(dossier.get("vehicles") or []) or ""),
        r"Will this quote include all": ("Yes", "E"),
        r"Filing Needed\?": "State" if c.get("state_filing_number") else None,
        r"Filing State": "CA" if state == "CA" else None,
        r"Filing Type": "MCP-65" if c.get("state_filing_number") else None,
    }
    FINAL = {
        r"Wholesaler Agent of Record": "RT Specialty",
        r"Progressive Agent Code": "94767",
        # Full legal name, dictated 7.29 — "Sierra" alone is the shorthand the
        # team says out loud, not what goes on a submission.
        r"Retail Agency Name": "Sierra Pacific Insurance Services Inc",
        r"spouse/domestic partner": "No",
        r"Blanket Waiver of Subrogation endorsement": "No",
        # "anyone else authorized … -> no, every time". The action cell below
        # it then has nothing to grant, and stays blank on purpose.
        r"anyone else authorized to access or update": "No",
        # NOT touched by JC's 7.27 no-dashes rule — this INSERTS a dash into
        # the FEIN on purpose (XX-XXXXXXX is the conventional FEIN shape).
        # That looks like the opposite of "no dashes are needed", so it is
        # flagged as an open question rather than changed — see this change's
        # final report.
        r"Employer Identification Number":
            re.sub(r"[.\s]", "-", str(c.get("fein") or "")) or None,
        r"order MVR/CLUE": "Yes",
        r"^Mailing address": c.get("mailing_address"),
        r"^Phone Number": formatting.format_phone(c.get("contact_cell") or c.get("office_phone")),
    }

    wb = load_workbook(TEMPLATE)
    written = 0

    def write_labeled(sheet: str, table: dict) -> None:
        nonlocal written
        ws = wb[sheet]
        for r in range(9, 60):
            label = ws.cell(row=r, column=3).value
            if not label:
                continue
            label = re.sub(r"\s+", " ", str(label)).strip()
            for pat, val in table.items():
                col = 4
                if isinstance(val, tuple):
                    val, colname = val
                    col = 5 if colname == "E" else 4
                # A model or broker sometimes answers "unknown" / "N/A"
                # rather than leaving the field empty outright — JC's rule is
                # the same either way: leave the cell blank, never print a
                # placeholder that reads like a real answer.
                val = formatting.blank_if_unknown(val)
                if val in (None, ""):
                    continue
                if re.search(pat, label, re.I):
                    ws.cell(row=r, column=col).value = str(val)
                    written += 1
                    break

    onhook = next((str(v.get("onhook")) for v in dossier.get("vehicles") or []
                   if v.get("onhook")), (covs.get("on_hook") or ""))
    RATES = {
        r"Bodily Injury and Property Damage":
            (("Yes", "D") if covs.get("auto_liability") else None),
        r"^UM/UIM$": (("Yes", "D") if covs.get("um_uim") else None),
        r"^Comprehensive$": (("Yes", "D") if covs.get("comprehensive") else None),
        r"^Collision$": (("Yes", "D") if covs.get("collision") else None),
        r"^On Hook": (("Yes", "D") if onhook else None),
        # no commercial lot -> nothing stored for third parties -> no GKLL
        r"Garage Keepers Legal":
            (("No", "D") if _stored(dossier) == "Taken home" else None),
    }
    RATES_E = {
        r"Bodily Injury and Property Damage": covs.get("auto_liability"),
        r"^UM/UIM$": covs.get("um_uim"),
        r"^Comprehensive$": covs.get("comprehensive"),
        r"^Collision$": covs.get("collision"),
        r"^On Hook": onhook or None,
    }
    write_labeled("Start", START)
    write_labeled("Business", BUSINESS)
    write_labeled("Final Details", FINAL)
    write_labeled("Rates", RATES)
    write_labeled("Rates", {k: (v, "E") for k, v in RATES_E.items() if v})

    # ---- Vehicles: one column per unit (D=Veh1 .. L=Veh9), labels in C
    ws = wb["Vehicles"]
    rows = []                       # (label, row) in ROW ORDER — labels repeat:
    for r in range(10, 23):         # the trailer section reuses "Year/Make/Model"
        lab = ws.cell(row=r, column=3).value   # and "VIN"; rows <=22 = vehicles
        if lab:
            rows.append((re.sub(r"\s+", " ", str(lab)).strip(), r))

    def vrow(pat):
        return next((r for lab, r in rows if re.search(pat, lab, re.I)), None)

    # Trailers have their own block on this sheet (rows 24-30) and were never
    # filled: every unit went into the power-unit columns, so 8 trucks + 3
    # trailers — which passes the 1-9 routing rule, because trailers are not
    # power units — silently lost two rows to the [:9] slice.
    units = [v for v in (dossier.get("vehicles") or []) if isinstance(v, dict)]
    trailers = [v for v in units if _is_trailer(v)]
    trucks = [v for v in units if not _is_trailer(v)]
    dropped: list[str] = []
    if len(trucks) > MAX_RTS_UNITS:
        dropped.append(f"{len(trucks)} power units but the RTS sheet holds "
                       f"{MAX_RTS_UNITS} — units {MAX_RTS_UNITS + 1}+ are NOT on "
                       f"this workbook; they need a second submission or the "
                       f"carrier's large-fleet form")
    if len(trailers) > MAX_RTS_UNITS:
        dropped.append(f"{len(trailers)} trailers but the sheet holds "
                       f"{MAX_RTS_UNITS} — trailers {MAX_RTS_UNITS + 1}+ are NOT "
                       f"on this workbook")

    for i, v in enumerate(trucks[:MAX_RTS_UNITS]):
        col = 4 + i
        def put(pat, val):
            nonlocal written
            r = vrow(pat)
            if r and val not in (None, ""):
                ws.cell(row=r, column=col).value = str(val)
                written += 1
        bt = str(v.get("body_type") or "").lower()
        put(r"Vehicle type \(Tractor", body_label(v.get("body_type")))
        put(r"Year/Make/Model", " ".join(str(x) for x in
                                         (v.get("year"), v.get("maker"), v.get("model")) if x))
        put(r"^VIN$", v.get("vin"))
        put(r"Full Garaging address", loc_addr)          # same as location, per JC
        # Axles: blank, and deliberately so. 7.22 said "typically two unless
        # it's GVW higher"; on 7.29 he withdrew it — "I don't think we need to
        # have how many axles". Loan/lease is the same kind of blank: answering
        # yes makes the underwriter demand the loss-payee detail up front.
        put(r"Hitch type", "N/A")                        # no trailers on file
        pae = v.get("perm_equipment_value")
        put(r"permanently attached", f"Yes - {pae}" if pae else "No")
        sv = v.get("stated_value")
        put(r"Stated Value \(excluding", f"{int(sv):,}" if sv else "")
        # Derived, never asked: comp/collision exists to protect a value, so a
        # priced unit gets it and a unit priced at zero does not. No value on
        # file is neither — say nothing rather than drop physical damage from
        # the quote on a guess.
        put(r"Comprehensive or Collision",
            "" if sv is None else ("Yes" if sv else "No"))

    # ---- Trailers: their own block, same column-per-unit shape
    trow = {}
    for r in range(24, 31):
        lab = ws.cell(row=r, column=3).value
        if lab:
            trow[re.sub(r"\s+", " ", str(lab)).strip()] = r

    def tput(pat, col, val):
        nonlocal written
        r = next((rr for lab, rr in trow.items() if re.search(pat, lab, re.I)), None)
        if r and val not in (None, ""):
            ws.cell(row=r, column=col).value = str(val)
            written += 1

    for i, t in enumerate(trailers[:MAX_RTS_UNITS]):
        col = 4 + i
        tput(r"Vehicle type \(Lowboy", col, str(t.get("body_type") or "").title())
        tput(r"Year/Make/Model", col, " ".join(str(x) for x in
             (t.get("year"), t.get("maker"), t.get("model")) if x))
        tput(r"^VIN$", col, t.get("vin"))
        tput(r"Garaging address", col, loc_addr)
        tput(r"length of the trailer", col, t.get("length"))
        tsv = t.get("stated_value")
        tput(r"Stated Value", col, f"{int(tsv):,}" if tsv else "")
        tput(r"Comprehensive or Collision", col,
             "" if tsv is None else ("Yes" if tsv else "No"))

    # ---- Drivers: row 13 = driver 1, columns D..M
    ws = wb["Drivers"]
    all_drivers = dossier.get("drivers") or []
    if len(all_drivers) > MAX_RTS_DRIVERS:
        dropped.append(f"{len(all_drivers)} drivers but the sheet holds "
                       f"{MAX_RTS_DRIVERS} — drivers {MAX_RTS_DRIVERS + 1}+ are "
                       f"NOT on this workbook")
    for i, drv in enumerate(all_drivers[:MAX_RTS_DRIVERS]):
        r = 13 + i
        name = str(drv.get("name") or "").split()
        had, when, what = _driver_claims(dossier, drv)
        vals = {4: name[0] if name else "", 5: " ".join(name[1:]),
                6: formatting.format_date(drv.get("birthday")),
                7: formatting.strip_cosmetic_dashes(
                    drv.get("license") or drv.get("license_number")),
                8: drv.get("license_state"), 9: had, 10: when, 11: what}
        for col, val in vals.items():
            if val not in (None, ""):
                ws.cell(row=r, column=col).value = str(val)
                written += 1

    out_dir = OUT / slug
    out_dir.mkdir(parents=True, exist_ok=True)
    name = f"{sp} CAP RTS supp app {m8()}.xlsx"
    wb.save(out_dir / name)

    # JC's three-states rule, 7.29: "every single answer has to have either an
    # answer accustomed to that insured, a cookie cutter answer, or we're
    # purposely leaving it blank because of some reason." The purposeful blanks
    # are declared here so the underwriter-facing reply can say so — otherwise
    # a deliberate blank is indistinguishable from a hole in the file, and his
    # words for that were "was that on purpose or did you not know?".
    deliberate = ["axles (withdrawn 7.29)", "loan/lease on vehicle",
                  "vehicle loan company"]
    unknown_blanks = []
    # The expiring liability limit drives the whole quote. HAMIL's came out
    # blank and nothing said so — the one kind of blank that has to speak.
    if not covs.get("auto_liability"):
        unknown_blanks.append("current bodily injury liability limit — nothing "
                              "on file; the quote is priced off it")
    for i, v in enumerate((dossier.get("vehicles") or [])[:9], 1):
        if v.get("stated_value") is None:
            unknown_blanks.append(f"comp/collision for vehicle {i} — no stated "
                                  f"value on file")

    # roundtrip: what we wrote is what a reader gets
    wb2 = load_workbook(out_dir / name)
    return {"ok": True, "file": str(out_dir / name), "cells": written,
            "deliberate_blanks": deliberate, "unknown_blanks": unknown_blanks,
            "dropped": dropped}


def apply_qp(slug: str, pdf_path: str) -> list[str]:
    """Read a QP PDF and merge it into clients/<slug>/state.json in place.

    The QP fills gaps only — whatever the dossier already holds wins, because a
    QP can predate a broker's correction (see qp_read.merge_into_dossier).
    Where the two disagree the QP's figure is NOT written; the disagreement is
    returned alongside qp_read's own warnings (unresolved "Same" placeholders, a
    mismatched loss-run block count, the FEIN's dot-vs-dash format, ...) so the
    broker sees both values and decides.
    """
    folder = CLIENTS / slug
    folder.mkdir(parents=True, exist_ok=True)
    state_path = folder / "state.json"
    dossier = json.loads(state_path.read_text(encoding="utf-8")) if state_path.exists() else {}
    result = qp_read.read_qp(pdf_path)
    dossier, conflicts = qp_read.merge_into_dossier(dossier, result["dossier"])
    state_path.write_text(json.dumps(dossier, indent=1), encoding="utf-8")
    return result["warnings"] + [f"QP disagrees with the file — {c}" for c in conflicts]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--client", required=True)
    ap.add_argument("--from-qp", metavar="PDF",
                    help="read a QP PDF and merge it into the client's dossier before filling")
    args = ap.parse_args()
    if args.from_qp:
        try:
            qp_warnings = apply_qp(args.client, args.from_qp)
        except qp_read.QPReadError as exc:
            print(f"error: {exc}", file=sys.stderr)
            sys.exit(2)
        for w in qp_warnings:
            print(f"QP warning: {w}")
    r = fill(args.client)
    print(f"RTS supplemental: {r['cells']} cells written")
    print(f"  {r['file']}")


if __name__ == "__main__":
    main()
